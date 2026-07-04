from openpyxl import load_workbook
wb=load_workbook('source/bearings.xlsx', data_only=True)
for name in wb.sheetnames:
    ws=wb[name]
    print('Sheet:', name, 'max_row:', ws.max_row, 'max_col:', ws.max_column)
    found=[]
    for r in range(1, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            v=ws.cell(row=r,column=c).value
            if v is None: continue
            if 'SUPER' in str(v).upper():
                found.append((r,c,str(v)))
    print('Found', len(found), 'matches')
    for f in found[:50]: print(f)
